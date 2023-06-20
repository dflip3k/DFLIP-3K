import pandas as pd
from openpyxl import load_workbook
import pickle


# 定义要读取的Excel文件路径和文件名
excel_files = ['AIGODLIKE0.xlsx', 'AIGODLIKE1.xlsx', 'AIGODLIKE2.xlsx', 'AIGODLIKE3.xlsx', 'AIGODLIKE4.xlsx', 'AIGODLIKE5.xlsx'
               , 'AIGODLIKE6.xlsx']

data = {}

for file in excel_files:
    workbook = load_workbook(filename=file)
    sheet = workbook.active

    for row in sheet.iter_rows(values_only=True):
        # import pdb;pdb.set_trace()

        rawurl = "https://ocdn.aigodlike.com/img/paintings/"+str(row[2].split('/')[-2].split('?')[0])
        if int(row[0]) not in data.keys():
            data[int(row[0])] = [row[1], row[2], rawurl]
            print(rawurl)

# import pdb;pdb.set_trace()
def cun(s):
    file = open(s,'wb')
    pickle.dump(data,file)
    file.close()

cun('./data')