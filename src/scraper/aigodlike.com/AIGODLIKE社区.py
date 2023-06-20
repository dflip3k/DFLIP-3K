
import requests
import json
import time
import re
from bs4 import BeautifulSoup as bs
import pickle
import openpyxl
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

def az(s):

    u =[]

    for i in range(ord("A"),ord("Z")+1):
        u.append(chr(i))
        if len(u)==s:
            return u
        else:
            pass

    for ii in range(ord("A"),ord("Z")+1):

        for i in range(ord("A"),ord("Z")+1):
            u.append(chr(ii)+chr(i))
            if len(u)==s:
                return u
            else:
                pass
def excel():

    wb = openpyxl.Workbook()
    word = wb.active
    
    for i in range(len(data)):

        a = az(len(data[i]))
        for r in range(len(a)):
            
            su = a[r]+str(i+1)
            word[su]=data[i][r]
            # print(data[i][r])

    a = '搜索结果'
    wb.save('{0}.xlsx'.format(a))
    print('已完成，已保存好了')
         
#===================================================

def hmtl_content(s):
    for i in range(1):
        headers ={}
        headers["authority"]='api.aigodlike.com'
        headers["method"]='GET'
        headers["path"]='/v1/content/illustrated-handbook/index?page=2&orderType=1&asc=false'
        headers["scheme"]='https'
        headers["accept"]='application/json, text/plain, */*'
        headers["accept-encoding"]='gzip, deflate, br'
        headers["accept-language"]='zh-CN,zh;q=0.9'
        headers["origin"]='https://www.aigodlike.com'
        headers["referer"]='https://www.aigodlike.com/'
        headers["sec-fetch-dest"]='empty'
        headers["sec-fetch-mode"]='cors'
        headers["sec-fetch-site"]='same-site'
        headers["user-agent"]='Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36'

        url ='https://api.aigodlike.com/v1/content/illustrated-handbook/index?page={0}&orderType=1&asc=false'.format(s)

        html  =requests.get(url,headers=headers)
        return html

def con_open(s):
    for i in range(1):
        headers ={}
        headers["authority"]='api.aigodlike.com'
        headers["method"]='GET'
        headers["path"]='/v1/content/illustrated-handbook/detail?illustratedHandbookId=1640650885094703104'
        headers["scheme"]='https'
        headers["accept"]='application/json, text/plain, */*'
        headers["accept-encoding"]='gzip, deflate, br'
        headers["accept-language"]='zh-CN,zh;q=0.9'
        headers["origin"]='https://www.aigodlike.com'
        headers["referer"]='https://www.aigodlike.com/paintings?id=1640650885094703104'
        headers["sec-fetch-dest"]='empty'
        headers["sec-fetch-mode"]='cors'
        headers["sec-fetch-site"]='same-site'
        headers["user-agent"]='Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36'

        url ='https://api.aigodlike.com/v1/content/illustrated-handbook/detail?illustratedHandbookId={0}'.format(s)

        html  =requests.get(url,headers=headers)
        return html

    
def main():

    for sss in range(0,72):
        try:
            html =hmtl_content(sss)

            con = json.loads(html.text)

            dic = con['data']

            x=[]
            for i in dic:
                img ='https://aigodlike-prod.oss-cn-beijing.aliyuncs.com/img/paintings/{0}?x-oss-process=style/preview'.format(i['imgIdf'])
                u = [i['id'],img]
                # print(u)
                x.append(u)

            for s in x:
                try:
                    html = con_open(s[0])

                    con = json.loads(html.text)

                    dic = con['data']['fullMeta']

                    data.append([s[0], dic, s[1]])
                    print(s[0], '已下载')
                except:
                    print('failed download')
                time.sleep(1)

            wb = openpyxl.Workbook()
            word = wb.active

            for i in range(len(data)):
                a = az(len(data[i]))
                for r in range(len(a)):
                    su = a[r] + str(i + 1)
                    try:
                        word[su] = data[i][r]
                    except:
                        word[su] = ILLEGAL_CHARACTERS_RE.sub(r'', data[i][r])
            a = '搜索结果'
            wb.save('{0}.xlsx'.format(a))
            print('已完成{}，已保存好了'.format(sss))
        except:
            print("error {}".format(sss))



data=[]

main()

excel()

            

            
