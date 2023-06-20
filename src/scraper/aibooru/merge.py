import pandas as pd
from openpyxl import load_workbook
import pickle

# 创建空字典，用于存储所有Excel文件中的数据
data_dict = {}

# 定义要读取的Excel文件路径和文件名
excel_files = ['1.xlsx', '2.xlsx']

# 读取Excel文件
data = {}

for file in excel_files:
    workbook = load_workbook(filename=file)
    sheet = workbook.active

    for row in sheet.iter_rows(values_only=True):
        # import pdb;pdb.set_trace()
        id = row[1].split('/')[-1]
        last_segment = row[0].split('/')[-3:]
        sub_string = last_segment[-1].split('.')[0]

        new_url = 'https://cdn.aibooru.online/original/{}/{}/{}.png'.format(last_segment[0], last_segment[1], sub_string)

        if int(id) not in data.keys():
            data[int(id)] = [row[1], row[0], new_url]

def cun(s):
    file = open(s,'wb')
    pickle.dump(data,file)
    file.close()

cun('data')
